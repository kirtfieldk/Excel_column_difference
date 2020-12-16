from Columns import columns
from openpyxl import load_workbook

"""
    Function will added columns to the excel document
"""


class Workbook_Worker():
    def __init__(self, workbook):
        self.wb = load_workbook(f"{workbook}.xlsx")
        self.sheet = self.wb.active
        self.columns_in_sheet = []
        self.unique_values = []
        self.rows = [x.split('\n') for x in columns]
        # self.add_column()
        """ Save the workbook """
        self.read_columns()
        self.compare_columns()

        # self.wb.save(f'{workbook}_new.xlsx')

    def add_column(self):
        i = 1
        col = 1
        for x in self.rows:
            if(len(x) > 1):
                for value in x:
                    self.sheet.cell(column=col, row=i).value = value
                    i += 1
            col += 1

    def read_columns(self):
        """
            We iterate through all the columns append A1,B1,... so we read every colummn
            we then append all the values of each column into another 2D array
            We iterate through the max row in the excel sheet, if a cell has no value
            it is None
        """
        letter = 'A'
        for index in range(0, len(columns)):
            self.columns_in_sheet.append([])
            for row in range(1, self.sheet.max_row+1):
                value = self.sheet[f"{letter}{row}"].value
                if(value != None):
                    self.columns_in_sheet[index].append(
                        value)
            letter = chr(ord(letter) + 1)

    def compare_columns(self):
        self.unique_values = []
        i = 0
        for col in self.columns_in_sheet:
            self.unique_values.append([])
            for value in col:
                for second_column in self.columns_in_sheet:
                    if(second_column != col):
                        found = Workbook_Worker.find_if_substring(
                            value, second_column)
                        if(not found):
                            self.unique_values[i].append(value)
            i += 1
        print(self.unique_values)

    @staticmethod
    def find_if_substring(value, column_array):
        if(len(value) < 3):
            return False
        for val in column_array:
            if(value in val):
                return True
            if(val in value):
                return True
        return False


if __name__ == '__main__':
    Workbook_Worker('numbers_new')
